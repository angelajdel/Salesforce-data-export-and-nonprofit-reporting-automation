#!/usr/bin/env python3
"""
build_excel_report.py
---------------------------------
Builds a polished, multi-tab Excel workbook from the CLEANED client and
donor CSVs (output of clean_salesforce_export.py).

Tabs created:
  1. Dashboard  - KPI summary with live formulas + a chart
  2. Clients    - full cleaned client list, formatted as a table
  3. Donors     - full cleaned donor list, formatted as a table
  4. Data Cleanup Log - what was removed and why (for transparency/audit)

USAGE
-----
    python build_excel_report.py \
        --clients outputs/cleaned_clients.csv \
        --donors outputs/cleaned_donors.csv \
        --clients-log outputs/clients_clean_log.json \
        --donors-log outputs/donors_clean_log.json \
        --output outputs/Monthly_Report.xlsx
"""

import argparse
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

NAVY = "1F3864"
GOLD = "C9972A"
LIGHT_GRAY = "F2F2F2"
WHITE_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=11)
TITLE_FONT = Font(color=NAVY, bold=True, name="Arial", size=18)
LABEL_FONT = Font(color="595959", name="Arial", size=10)
KPI_FONT = Font(color=NAVY, bold=True, name="Arial", size=22)
BODY_FONT = Font(name="Arial", size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_table_sheet(ws, df, sheet_title):
    ws.sheet_view.showGridLines = False
    ws["A1"] = sheet_title
    ws["A1"].font = TITLE_FONT
    ws.append([])
    start_row = 3
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = WHITE_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    last_row = start_row + len(df)
    last_col = len(df.columns)
    for j in range(1, last_col + 1):
        col_letter = get_column_letter(j)
        max_len = max(
            [len(str(df.columns[j - 1]))]
            + [len(str(v)) for v in df.iloc[:, j - 1].astype(str)]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    table_ref = f"A{start_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=f"Tbl_{sheet_title.replace(' ', '')}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=False
    )
    ws.add_table(table)
    return start_row, last_row


def build_dashboard(wb, clients_df, donors_df, clients_log, donors_log):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Monthly Impact Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = "Auto-generated from cleaned Salesforce export"
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:F2")

    kpi_cells = [
        ("Active Clients", "=COUNTIF(Clients!H:H,\"Active\")", "B5"),
        ("Total Clients (all statuses)", "=COUNTA(Clients!A4:A10000)", "D5"),
        ("Total Raised This Period", "=SUM(Donors!D:D)", "B10"),
        ("Number of Donors", "=COUNTA(Donors!A4:A10000)", "D10"),
        ("Test Records Removed", clients_log["test_records_removed"] + donors_log["test_records_removed"], "B15"),
        ("Duplicates Removed", clients_log["duplicate_records_removed"] + donors_log["duplicate_records_removed"], "D15"),
    ]

    for label, formula, cell in kpi_cells:
        col = cell[0]
        row = int(cell[1:])
        label_cell = ws.cell(row=row, column=ws[cell].column, value=label)
        label_cell.font = LABEL_FONT
        val_cell = ws.cell(row=row + 1, column=ws[cell].column, value=formula)
        val_cell.font = KPI_FONT

    # Program breakdown table (feeds a chart) placed at row 20
    ws["A20"] = "Clients by Program"
    ws["A20"].font = Font(bold=True, name="Arial", size=12, color=NAVY)
    programs = sorted(clients_df["Program"].unique())
    ws.cell(row=21, column=1, value="Program").font = WHITE_FONT
    ws.cell(row=21, column=1).fill = HEADER_FILL
    ws.cell(row=21, column=2, value="Client Count").font = WHITE_FONT
    ws.cell(row=21, column=2).fill = HEADER_FILL
    for i, prog in enumerate(programs, start=22):
        ws.cell(row=i, column=1, value=prog).font = BODY_FONT
        ws.cell(row=i, column=2, value=f'=COUNTIF(Clients!F:F,"{prog}")').font = BODY_FONT

    chart = BarChart()
    chart.title = "Clients by Program"
    chart.y_axis.title = "Clients"
    chart.style = 10
    data = Reference(ws, min_col=2, min_row=21, max_row=21 + len(programs))
    cats = Reference(ws, min_col=1, min_row=22, max_row=21 + len(programs))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 8
    ws.add_chart(chart, "D20")

    # Donations by campaign
    ws["A32"] = "Donations by Campaign"
    ws["A32"].font = Font(bold=True, name="Arial", size=12, color=NAVY)
    campaigns = sorted(donors_df["Campaign"].unique())
    ws.cell(row=33, column=1, value="Campaign").font = WHITE_FONT
    ws.cell(row=33, column=1).fill = HEADER_FILL
    ws.cell(row=33, column=2, value="Total Raised").font = WHITE_FONT
    ws.cell(row=33, column=2).fill = HEADER_FILL
    for i, camp in enumerate(campaigns, start=34):
        ws.cell(row=i, column=1, value=camp).font = BODY_FONT
        ws.cell(row=i, column=2, value=f'=SUMIF(Donors!F:F,"{camp}",Donors!D:D)').font = BODY_FONT
        ws.cell(row=i, column=2).number_format = "$#,##0"

    pie = PieChart()
    pie.title = "Donations by Campaign"
    data = Reference(ws, min_col=2, min_row=33, max_row=33 + len(campaigns))
    cats = Reference(ws, min_col=1, min_row=34, max_row=33 + len(campaigns))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.width = 14
    pie.height = 8
    ws.add_chart(pie, "D32")

    for col, width in zip("ABCDEF", [26, 20, 4, 26, 20, 4]):
        ws.column_dimensions[col].width = width


def build_cleanup_log_sheet(wb, clients_log, donors_log):
    ws = wb.create_sheet("Data Cleanup Log")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Data Cleanup Log"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Records excluded from this report and why"
    ws["A2"].font = LABEL_FONT

    row = 4
    for label, log in [("CLIENTS", clients_log), ("DONORS", donors_log)]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color=NAVY, name="Arial", size=12)
        row += 1
        ws.cell(row=row, column=1, value="Total rows in export").font = LABEL_FONT
        ws.cell(row=row, column=2, value=log["total_rows_in_export"]).font = BODY_FONT
        row += 1
        ws.cell(row=row, column=1, value="Test records removed (ID contains 'TEST')").font = LABEL_FONT
        ws.cell(row=row, column=2, value=log["test_records_removed"]).font = BODY_FONT
        row += 1
        ws.cell(row=row, column=1, value="Duplicate records removed").font = LABEL_FONT
        ws.cell(row=row, column=2, value=log["duplicate_records_removed"]).font = BODY_FONT
        row += 1
        ws.cell(row=row, column=1, value="Clean rows remaining").font = LABEL_FONT
        ws.cell(row=row, column=2, value=log["clean_rows_remaining"]).font = BODY_FONT
        row += 2

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20


def main():
    parser = argparse.ArgumentParser(description="Build the monthly Excel report.")
    parser.add_argument("--clients", required=True)
    parser.add_argument("--donors", required=True)
    parser.add_argument("--clients-log", required=True)
    parser.add_argument("--donors-log", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    clients_df = pd.read_csv(args.clients, dtype=str, keep_default_na=False)
    donors_df = pd.read_csv(args.donors, dtype=str, keep_default_na=False)
    donors_df["Amount"] = pd.to_numeric(donors_df["Amount"], errors="coerce")

    with open(args.clients_log) as f:
        clients_log = json.load(f)
    with open(args.donors_log) as f:
        donors_log = json.load(f)

    wb = Workbook()
    build_dashboard(wb, clients_df, donors_df, clients_log, donors_log)

    ws_clients = wb.create_sheet("Clients")
    style_table_sheet(ws_clients, clients_df, "Client Roster")

    ws_donors = wb.create_sheet("Donors")
    donors_df_display = donors_df.copy()
    style_table_sheet(ws_donors, donors_df_display, "Donor Log")
    for row in ws_donors.iter_rows(min_row=4, min_col=4, max_col=4):
        for c in row:
            c.number_format = "$#,##0"

    build_cleanup_log_sheet(wb, clients_log, donors_log)

    wb.save(args.output)
    print(f"✔ Workbook written to {args.output}")


if __name__ == "__main__":
    main()
