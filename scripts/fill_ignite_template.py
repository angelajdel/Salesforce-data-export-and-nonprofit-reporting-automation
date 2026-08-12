#!/usr/bin/env python3
"""
fill_ignite_template.py
---------------------------------
Fills ONE month's column of the real "Ignite FY26 Outcomes - Template.xlsx"
with the outcome numbers Impact Hub computes, and leaves every other cell in
the workbook exactly as it already is -- same formatting, same labels, same
185+ rows Impact Hub doesn't (yet) compute.

WHY A SCRIPT AND NOT JUST A BROWSER EXPORT
--------------------------------------------
Your template has exact fonts, column widths, and cell-level number formats
(Times New Roman, "0%" formats already applied to every percentage cell,
etc.) built in. Rebuilding that faithfully from a browser tool would either
take real effort to match pixel-for-pixel or risk drifting from your actual
file. Instead, this script opens your REAL template and writes only into the
cells that match, so nothing about your existing file changes except the
values Impact Hub actually knows.

WHAT GETS FILLED
------------------
Impact Hub currently computes:
  - Education and Employment (30+) domain            -> "Ignite Domains of Care"
  - Reasons for Homelessness (by keyword match)        -> "Reasons for Homelessness"
  - Total clients served, active clients per program   -> "Program Counts"
  - Utilization % (overall + per program)              -> "Utilization"

Everything else in the template (New Enrollments, per-program exit/outcome
detail, Street Outreach, Residential Therapy, etc.) is left exactly as it
was in your file -- Impact Hub doesn't have that data yet, so this script
does not guess at it or blank it out.

Matching is done by SECTION first (using the template's own section header
rows, e.g. "Program Counts", "Utilization") and then by keyword inside that
section, so it won't accidentally write into the similarly-named
Bronzeville-specific "Percentage of BYS youth identify family conflict..."
row when filling the org-wide "Reasons for Homelessness" section.

USAGE
-----
    python fill_ignite_template.py \
        --template "Ignite_FY26_Outcomes_-_Template.xlsx" \
        --data outcomes_export.xlsx \
        --month Aug \
        --output "Ignite_FY26_Outcomes_-_Filled.xlsx"

`outcomes_export.xlsx` is downloaded from Impact Hub's Import step
("Export for the Ignite FY26 Outcomes Template") -- it's a small, plain
workbook you can also just open and read directly; this script is only
needed to merge those numbers into your real, fully-formatted template.

Run this again each month with that month's fresh export -- it always
starts from your real template (or from last month's filled copy, if you
pass that in as --template) and only touches the one month's column.
"""

import argparse
import re
import sys

from openpyxl import load_workbook


def extract_reason_phrase(label):
    """Pulls the specific concept out of a template row like
    'Percentage of youth identify family conflict as a presenting issue'
    -> 'family conflict'. Returns None if the row doesn't match that pattern."""
    m = re.search(r"identify (.+?) as a presenting issue", label, re.IGNORECASE)
    return m.group(1).strip().lower() if m else None


def find_reason_row(ws, row_range, reason_name):
    """Matches a configured Reason name to a template row by comparing the
    row's specific concept phrase (e.g. 'family conflict') against the
    Reason name -- not by splitting into individual words, which false-matches
    on common words like 'of' or 'a'."""
    name_lower = reason_name.lower().strip()
    for row in row_range:
        label = ws.cell(row=row, column=1).value or ""
        phrase = extract_reason_phrase(label)
        if not phrase:
            continue
        if phrase in name_lower or name_lower in phrase:
            return row
    return None

MONTH_COLUMNS = {
    "Jul": "B", "Aug": "C", "Sep": "D",
    "Oct": "F", "Nov": "G", "Dec": "H",
    "Jan": "J", "Feb": "K", "Mar": "L",
    "Apr": "N", "May": "O", "June": "P",
}


def find_sections(ws):
    """Returns a list of (start_row, end_row, section_name) using the
    template's own section-header rows (identified by column B == 'Jul')."""
    headers = []
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == "Jul":
            headers.append((row, ws.cell(row=row, column=1).value))
    sections = []
    for i, (row, name) in enumerate(headers):
        end = headers[i + 1][0] - 1 if i + 1 < len(headers) else ws.max_row
        sections.append((row + 1, end, name))
    return sections


def rows_in_section(ws, sections, section_name):
    for start, end, name in sections:
        if name and name.strip().lower() == section_name.strip().lower():
            return range(start, end + 1)
    return range(0, 0)


def find_row(ws, row_range, keyword_all=None, keyword_any=None):
    """Finds the first row (within row_range) whose column-A label matches.
    keyword_all: every keyword must appear (case-insensitive).
    keyword_any: at least one keyword must appear."""
    for row in row_range:
        label = (ws.cell(row=row, column=1).value or "").lower()
        if keyword_all and not all(k.lower() in label for k in keyword_all):
            continue
        if keyword_any and not any(k.lower() in label for k in keyword_any):
            continue
        if keyword_all or keyword_any:
            return row
    return None


def parse_outcomes_export(path):
    """Reads the small outcomes_export.xlsx that Impact Hub downloads and
    turns it back into the same data shape the fill logic expects. The
    export uses plain marker rows (e.g. 'ACTIVE_BY_PROGRAM') in column A
    followed by a two-column mini-table -- this walks the sheet once and
    collects everything between markers."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    data = {"activeByProgram": {}, "reasons": [], "utilizationByProgram": []}

    i = 0
    while i < len(rows):
        label = rows[i][0] if rows[i] else None

        if label == "TOTAL_CLIENTS_SERVED":
            data["totalClientsServed"] = rows[i][1]
        elif label == "EDUCATION_DOMAIN_PCT":
            data["educationDomainPct"] = rows[i][1]
        elif label == "UTILIZATION_OVERALL_PCT":
            data["utilizationOverallPct"] = rows[i][1]
        elif label == "ACTIVE_BY_PROGRAM":
            i += 2  # skip marker row + its "Program / Active Count" header row
            while i < len(rows) and rows[i] and rows[i][0]:
                data["activeByProgram"][rows[i][0]] = rows[i][1]
                i += 1
            continue
        elif label == "REASONS":
            i += 2
            while i < len(rows) and rows[i] and rows[i][0]:
                data["reasons"].append({"name": rows[i][0], "pct": rows[i][1]})
                i += 1
            continue
        elif label == "UTILIZATION_BY_PROGRAM":
            i += 2
            while i < len(rows) and rows[i] and rows[i][0]:
                data["utilizationByProgram"].append({"program": rows[i][0], "pct": rows[i][1]})
                i += 1
            continue

        i += 1

    return data


def main():
    parser = argparse.ArgumentParser(description="Fill one month of the Ignite FY26 template from Impact Hub data.")
    parser.add_argument("--template", required=True, help="Path to the Ignite FY26 Outcomes template (or last month's filled copy)")
    parser.add_argument("--data", required=True, help="Path to outcomes_export.xlsx from Impact Hub")
    parser.add_argument("--month", required=True, choices=list(MONTH_COLUMNS.keys()), help="Which month column to fill")
    parser.add_argument("--output", required=True, help="Path to write the filled workbook")
    args = parser.parse_args()

    data = parse_outcomes_export(args.data)
    col = MONTH_COLUMNS[args.month]

    wb = load_workbook(args.template, data_only=False)
    ws = wb["FY26 Template"]
    sections = find_sections(ws)

    filled = []
    skipped = []

    # --- Ignite Domains of Care: Education and Employment (30+) ---
    domain_rows = rows_in_section(ws, sections, "Ignite Domains of Care")
    row = find_row(ws, domain_rows, keyword_all=["education and employment"])
    if row and data.get("educationDomainPct") is not None:
        cell = ws[f"{col}{row}"]
        cell.value = data["educationDomainPct"] / 100
        cell.number_format = "0%"
        filled.append(f"Row {row} (Education and Employment domain)")
    elif row:
        skipped.append(f"Row {row} (Education and Employment domain) -- no data provided")

    # --- Reasons for Homelessness ---
    reason_rows = rows_in_section(ws, sections, "Reasons for Homelessness")
    for reason in data.get("reasons", []):
        row = find_reason_row(ws, reason_rows, reason["name"])
        if row:
            cell = ws[f"{col}{row}"]
            cell.value = reason["pct"] / 100
            cell.number_format = "0%"
            filled.append(f"Row {row} (Reason: {reason['name']})")
        else:
            skipped.append(f"Reason '{reason['name']}' -- no matching template row found")

    # --- Program Counts ---
    count_rows = rows_in_section(ws, sections, "Program Counts")
    if data.get("totalClientsServed") is not None:
        row = find_row(ws, count_rows, keyword_all=["youth served by ignite"])
        if row:
            ws[f"{col}{row}"] = data["totalClientsServed"]
            filled.append(f"Row {row} (Total youth served)")

    residential_total = 0
    for program, count in data.get("activeByProgram", {}).items():
        row = find_row(ws, count_rows, keyword_all=["active in", program.lower()])
        if row:
            ws[f"{col}{row}"] = count
            filled.append(f"Row {row} (Active in {program})")
            residential_total += count
        else:
            skipped.append(f"Active count for '{program}' -- no matching template row found")

    if residential_total:
        row = find_row(ws, count_rows, keyword_all=["active in residential programming"])
        if row:
            ws[f"{col}{row}"] = residential_total
            filled.append(f"Row {row} (Active in Residential Programming -- summed)")

    # --- Utilization ---
    util_rows = rows_in_section(ws, sections, "Utilization")
    if data.get("utilizationOverallPct") is not None:
        row = find_row(ws, util_rows, keyword_all=["residential - bed utilization rate"])
        if row:
            cell = ws[f"{col}{row}"]
            cell.value = data["utilizationOverallPct"] / 100
            cell.number_format = "0%"
            filled.append(f"Row {row} (Overall Residential Bed Utilization)")

    for entry in data.get("utilizationByProgram", []):
        program_lower = entry["program"].lower()
        row = find_row(ws, util_rows, keyword_all=[program_lower, "utilization rate"])
        if row:
            cell = ws[f"{col}{row}"]
            cell.value = entry["pct"] / 100
            cell.number_format = "0%"
            filled.append(f"Row {row} ({entry['program']} Utilization Rate)")
        else:
            skipped.append(f"Utilization for '{entry['program']}' -- no matching template row found (HUD-funded / no-BYS variants are never auto-filled)")

    wb.save(args.output)

    print(f"Filled {args.month} column ({col}) in {args.output}\n")
    print(f"Filled {len(filled)} cell(s):")
    for f in filled:
        print(f"  \u2713 {f}")
    if skipped:
        print(f"\nSkipped {len(skipped)} item(s) (no data or no matching row):")
        for s in skipped:
            print(f"  - {s}")
    print(
        "\nEverything else in the template -- New Enrollments, per-program exit "
        "detail, Street Outreach, Residential Therapy, and all quarterly/FY26 "
        "rollup columns -- was left exactly as it was in the source file."
    )


if __name__ == "__main__":
    main()
